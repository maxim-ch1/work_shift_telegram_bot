import openpyxl
from datetime import datetime, timedelta
import calendar

wb = openpyxl.load_workbook(filename="Тест.xlsx")

month = datetime.strftime(datetime.now(), "%m")
day = datetime.strftime(datetime.now(), "%d")
days_how_much = calendar.monthrange(2021, int(month))[1]

week_list = [
    'Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс'
]

second_shift = []
time_off_week = []
time_off_day = []

def check_sheet(switch, param_m, param_d):

    counter_day = 0 #счетчик недели
    month = param_m
    day = param_d

    def month_now(a):

        months_list = [
            'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 
             'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
        ]

        ws = wb[
            str(months_list[int(month) - a])
        ] #KeyError если отсутствует следующий месяц

        return ws

    #функция получает дату и выдает день и наименование недели из соответствующего стобца

    def day_date(a):

        date = ''
        day = ''

        for i in range(int(a) + 1, int(a) + 2):
            day = str(month_now(1).cell(row=2, column=i).value)
            date = str(month_now(1).cell(row=1, column=i).value) + '.' + month

        return day, date

    while counter_day <= switch:  # проход по дням

        if day_date(int(day) + counter_day)[0] == 'None':
            pass
        else:
            second_shift.append('____________' + day_date(int(day) + counter_day)[0] +
                                '('+ day_date(int(day) + counter_day)[1] + ')____________')

            for i in range(3, 20):  # проход по линиям за дату. Если сотрудников больше 20,
                v = [month_now(1).cell(row=i, column=1 + int(day) + counter_day).value]
                if v != [None]:  # проверяет есть ли в строке что-то или нет. в случае None пасует.
                    if v[0] != 'О':
                        second_shift.append(str(month_now(1).cell(row=i, column=1).value))
                    elif v[0] == 'О':
                        time_off_day.append(str(month_now(1).cell(row=i, column=1).value))
                        time_off_week.append(str(month_now(1).cell(row=i, column=1).value) + ' - '
                            + day_date(int(day) + counter_day)[0] + ' ('+ day_date(int(day) + counter_day)[1] + ')')
                    else:
                        print('elsee')

            if day_date(int(day) + counter_day)[0] == 'ВС':  # если видит ВС то прерывает выполнение операции
                break
            else:
                pass

        counter_day = counter_day + 1

def weeks():

    second_shift.clear()
    time_off_week.clear()

    check_sunday = 0

    month = datetime.strftime(datetime.now() + timedelta(days=0), "%m")
    day = datetime.strftime(datetime.now() + timedelta(days=0), "%d")

    next_month = datetime.strftime(
        datetime.now() + timedelta(days=days_how_much - int(day) + 1), "%m")
    next_day = datetime.strftime(
        datetime.now() + timedelta(days=days_how_much - int(day) + 1), "%d")

    check_sheet(6, month, day)

    for i in second_shift:
        if '____________ВС' in i:
            check_sunday =+ 1

    if check_sunday != 1:
        print('yes')
        check_sheet(6, next_month, next_day)

    x = 'Дежурные на текущей неделе:\n'

    for i in second_shift:
        x = x + '\n' + i

    if time_off_week != []:
        x += ("\n\nОтгулы:\n")
        for i in time_off_week:
            x = x + '\n' + i
    return x

#print(weeks())

def next_week():

    second_shift.clear()
    time_off_week.clear()

    check_sunday = 0
    counter_day = datetime.today().weekday()

    month = datetime.strftime(datetime.now() + timedelta(days=7 - counter_day), "%m")
    day = datetime.strftime(datetime.now() + timedelta(days=7 - counter_day), "%d")

    next_month = datetime.strftime(
        datetime.now() + timedelta(days=days_how_much - int(day) + 8 - counter_day), "%m")
    next_day = datetime.strftime(
            datetime.now() + timedelta(days=days_how_much - int(day) + 8 - counter_day), "%d")

    check_sheet(6, month, day)

    for i in second_shift:
        if '____________ВС' in i:
            check_sunday =+ 1

    if check_sunday != 1:
        check_sheet(6, next_month, next_day)

    else:
        pass

    x = 'Дежурные на следующей неделе:\n'

    for i in second_shift:
        x = x + '\n' + i

    if time_off_week != []:
        x += ("\n\nОтгулы:\n")
        for i in time_off_week:
            x = x + '\n' + i

    return x

#print(next_week())

def day_yesterday():

    second_shift.clear()
    time_off_day.clear()
    print(datetime.strftime(datetime.now() - timedelta(days=1), "%m"))
    print(datetime.strftime(datetime.now() - timedelta(days=1), "%d"))

    check_sheet(0, datetime.strftime(datetime.now() - timedelta(days=1), "%m"),
           datetime.strftime(datetime.now() - timedelta(days=1), "%d"))

    x = 'Дежурные вчера:\n'

    for i in second_shift:
        x = x + '\n' + i

    if time_off_day != []:
        x += ("\n\nОтгулы:")
        for i in time_off_day:
            x = x + '\n' + i

    return x

#print(day_yesterday())

def day_today():

    second_shift.clear()
    time_off_day.clear()

    check_sheet(0, datetime.strftime(datetime.now(), "%m"), datetime.strftime(datetime.now(), "%d"))

    if day != days_how_much:
        pass
    else:
        if days_how_much - int(day) <= 2:
            check_sheet(0, month, day)
        else:
            pass

    x = 'Дежурные сегодня:\n'

    for i in second_shift:
        x = x + '\n' + i

    if time_off_day != []:
        x += ("\n\nОтгулы:")
        for i in time_off_day:
            x = x + '\n' + i

    return x

#print(day_today())

def day_tomorrow():

    second_shift.clear()
    time_off_day.clear()

    check_sheet(0, datetime.strftime(datetime.now() + timedelta(days=1), "%m"),
           datetime.strftime(datetime.now() + timedelta(days=1), "%d"))

    x = 'Дежурные на завтра:\n'

    for i in second_shift:
        x = x + '\n' + i

    if time_off_day != []:
        x += ("\n\nОтгулы:")
        for i in time_off_day:
            x = x + '\n' + i

    return x

#next_week()
#print(day_tomorrow())
